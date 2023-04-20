import openpyxl
import subprocess
from flask import render_template, request, jsonify, flash
import oracledb
import pandas as pd
from datetime import datetime
from app import app


@app.route('/')
def index():
    flash(f'Выберите тип отчёта')
    return render_template('index.html')


@app.route('/inner-page')
def inner_page():
    return render_template('inner-page.html')


@app.route('/weekly-report')
def weekly_report():
    return render_template('weekly-report.html')


@app.route('/complications')
def complications():
    flash('Страница на разработке')
    return render_template('complications.html')


@app.route('/', methods=['POST'])
def weekly():
    print('START')
    global date1
    global date2
    global full_year
    global week
    global date_begin_year
    global age_from
    global age_to
    global date_to_day
    global date_to_month
    global date_to_year
    global date_from_day
    global date_from_month
    global date_from_year
    global constant
    global date_from
    global date_to

    # data = request.form.to_dict(flat=False)
    # return jsonify(data)
    age_from = request.form['age_from']
    age_to = request.form['age_to']
    constant = request.form['constant']

    # block date:
    #date_from = DateForm.entrydate_from
    #date_to = DateForm.entrydate_to
    date_from = request.form['date_from']
    date_to = request.form['date_to']

    date_from_splited, date_to_splited = date_from.split('-'), date_to.split('-')
    try:
        date_from_day, date_from_month, date_from_year = date_from_splited[2], date_from_splited[1], \
                                                         date_from_splited[0]
    except IndexError:
        flash('Нужно выбрать диапазон дат')
        return render_template('weekly-report.html')

    date_to_day, date_to_month, date_to_year = date_to_splited[2], date_to_splited[1], date_to_splited[0]
    dt = datetime(int(date_from_year), int(date_from_month), int(date_from_day))
    week, full_year = dt.strftime("%W"), date_from_year
    date_begin_year = f'01.01.{date_from_year}'
    date1, date2 = f'{date_from_day}.{date_from_month}.{date_from_year}', f'{date_to_day}.{date_to_month}.' \
                                                                          f'{date_to_year}'
    print(f'Неделя: {week}\nВыбранные даты: с {date_from} по {date_to}')

    if request.method == 'POST':
        if request.form['submit'] == "Сформировать еженедельный отчёт":
            message = f"Выбранные даты: с {date_from} по {date_to}"
            flash(message)
            output_xlsx_flat()
            save_info_week(wb_obj_out_weekly)
        elif request.form['submit'] == "Ежемесячный":
            flash('На разработке')
            message = f"Выбранные даты: с {date_from} по {date_to}"
            flash(message)
            output_xlsx_flat()
            save_info_month(wb_obj_out_monthly)
        else:
            return 'Ошибка 01'
    return render_template('process_data.html')


def check_text_row(string, sub_str):
    if (string.find(sub_str) == -1):
        return False
    else:
        return True


def intersection_list(list1, list2):
   return set(list1).intersection(list2)


def two_in_one_cell(column_1, column_2, list_1, list_2, text_1, text_2, text_3='!@#'):
    for i, row in enumerate(column_1):
        if row.value == text_1:
            list_1.append(i)
    for i, row in enumerate(column_2):
        if row.value == text_2 or row.value == text_3:
            list_2.append(i)
    var = len(intersection_list(list_1, list_2))
    return var


def output_xlsx_flat():

    global sheet_obj_out
    global sheet_obj_out_m
    global wb_obj_out_weekly
    global wb_obj_out_monthly

    # инициализируем переменные
    summ_1g, summ_1g_1, summ_1g_2, summ_2g, summ_2g_1, summ_2g_2, summ_3g, summ_4g, summ_4g_1, summ_4g_2, summ_4g_3, \
    summ_4g_4, summ_5g, summ_6g, summ_6g_1, summ_7g, summ_7g_1, summ_7g_2, summ_7g_5, summ_7g_4, summ_8g, summ_8g_1, \
    summ_8g_2, summ_8g_3, summ_8g_4, summ_9g, summ_10g, summ_11g, summ_11g_1, summ_11g_2, summ_11g_3, summ_11g_4, \
    summ_11g_5, summ_11g_6, summ_11g_7, summ_11g_8, summ_11g_9, summ_12g, summ_12g_1, summ_vmp, summ_oms, \
    summ_pd, uniq_pers_deti, summ_aproba, count_sources, total_konserv, summ_foreign, summ_vmp_oms_from_begin_year, \
    summ_pleopto_kids_int_c006, summ_13g, summ_1g_previous_m_year = (0 for _ in range(51))
    #summ_13g = -1

    # формируем книги
    save_selected_summary_flat_data()
    save_selected_summary_sources_fin()
    save_selected_number_of_children()
    save_selected_summary_vmp()

    # открываем книги
    wb_obj = openpyxl.load_workbook(path_from)
    wb_obj_sources = openpyxl.load_workbook(path_sources)
    wb_obj_kids = openpyxl.load_workbook(path_kids)
    wb_obj_vmp = openpyxl.load_workbook(path_vmp)
    wb_obj_out_weekly = openpyxl.load_workbook(path_output_weekly)
    wb_obj_out_monthly = openpyxl.load_workbook(path_output_monthly)

    if request.method == 'POST':
        if request.form['submit'] == "Сформировать еженедельный отчёт":
            sheet_obj_out = wb_obj_out_weekly.active
        elif request.form['submit'] == "Ежемесячный":
            sheet_obj_out = wb_obj_out_monthly.active
            # новый блок для ежемесячного отчёта:



    sheet_obj = wb_obj.active
    sheet_obj_sources = wb_obj_sources.active
    sheet_obj_kids = wb_obj_kids.active
    sheet_obj_vmp = wb_obj_vmp.active

    column_e = sheet_obj['E']
    column_d = sheet_obj['D']
    column_v = sheet_obj['V']
    column_m = sheet_obj['M']
    column_n = sheet_obj['N']
    column_k = sheet_obj['K']
    column_g = sheet_obj['G']
    column_p = sheet_obj['P']
    column_f = sheet_obj['F']
    column_aj = sheet_obj['AJ']

    column_source_g = sheet_obj_sources['G']
    column_source_f = sheet_obj_sources['F']

    #column_kids_d = sheet_obj_sources['D']
    column_kids_d = sheet_obj_kids['C3'].value
    column_vmp_b = sheet_obj_vmp['B']

    #a, b = sheet_obj_out.max_row, sheet_obj_out.max_column


    if request.method == 'POST':
        if request.form['submit'] == "Сформировать еженедельный отчёт":
            calculating_excel(column_d, column_v, column_m, column_f, column_k, column_p, column_e, column_n, column_g,
                              column_aj, column_source_g, column_source_f, column_kids_d, column_vmp_b,
                              summ_1g, summ_1g_1, summ_1g_2, summ_2g, summ_2g_1, summ_2g_2, summ_3g, summ_4g, summ_4g_1,
                              summ_4g_2, summ_4g_3, summ_4g_4, summ_5g, summ_6g, summ_6g_1, summ_7g, summ_7g_1, summ_7g_2,
                              summ_7g_5, summ_7g_4, summ_8g, summ_8g_1, summ_8g_2, summ_8g_3, summ_8g_4, summ_9g, summ_10g,
                              summ_11g, summ_11g_1, summ_11g_2, summ_11g_3, summ_11g_4, summ_11g_5, summ_11g_6, summ_11g_7,
                              summ_11g_8, summ_11g_9, summ_12g, summ_12g_1, summ_vmp, summ_oms, summ_pd, uniq_pers_deti,
                              summ_aproba, count_sources, total_konserv, summ_foreign, summ_vmp_oms_from_begin_year,
                              summ_pleopto_kids_int_c006, summ_13g)
        elif request.form['submit'] == "Ежемесячный":
            calculating_excel(column_d, column_v, column_m, column_f, column_k, column_p, column_e, column_n, column_g,
                              column_aj, column_source_g, column_source_f, column_kids_d, column_vmp_b,
                              summ_1g, summ_1g_1, summ_1g_2, summ_2g, summ_2g_1, summ_2g_2, summ_3g, summ_4g, summ_4g_1,
                              summ_4g_2, summ_4g_3, summ_4g_4, summ_5g, summ_6g, summ_6g_1, summ_7g, summ_7g_1, summ_7g_2,
                              summ_7g_5, summ_7g_4, summ_8g, summ_8g_1, summ_8g_2, summ_8g_3, summ_8g_4, summ_9g, summ_10g,
                              summ_11g, summ_11g_1, summ_11g_2, summ_11g_3, summ_11g_4, summ_11g_5, summ_11g_6, summ_11g_7,
                              summ_11g_8, summ_11g_9, summ_12g, summ_12g_1, summ_vmp, summ_oms, summ_pd, uniq_pers_deti,
                              summ_aproba, count_sources, total_konserv, summ_foreign, summ_vmp_oms_from_begin_year,
                              summ_pleopto_kids_int_c006, summ_13g, summ_1g_previous_m_year)

            #новое, покопаться тут, по логике здесь дату - год, и инициализация книг

            date_from_year_old = int(date_from_year) - 1
            date_to_year_old = int(date_from_year) - 1
            date3, date4 = f'{date_from_day}.{date_from_month}.{date_from_year_old}', f'{date_to_day}.{date_to_month}.' \
                                                                                      f'{date_to_year_old}'
            date1, date2 = date3, date4
            save_selected_summary_flat_data()
            save_selected_summary_sources_fin()
            save_selected_number_of_children()
            save_selected_summary_vmp()
            print('DATA1 FLAG', date1)
            print('DATA2 FLAG', date2)

            wb_obj_old_year = openpyxl.load_workbook(path_from_old_year)
            wb_obj_sources_old_year = openpyxl.load_workbook(path_sources_old_year)
            wb_obj_kids_old_year = openpyxl.load_workbook(path_kids_old_year)
            wb_obj_vmp_old_year = openpyxl.load_workbook(path_vmp_old_year)
            sheet_obj_old_year = wb_obj_old_year.active
            sheet_obj_sources_old_year = wb_obj_sources_old_year.active
            sheet_obj_kids_old_year = wb_obj_kids_old_year.active
            sheet_obj_vmp_old_year = wb_obj_vmp_old_year.active

            column_e = sheet_obj_old_year['E']
            column_d = sheet_obj_old_year['D']
            column_v = sheet_obj_old_year['V']
            column_m = sheet_obj_old_year['M']
            column_n = sheet_obj_old_year['N']
            column_k = sheet_obj_old_year['K']
            column_g = sheet_obj_old_year['G']
            column_p = sheet_obj_old_year['P']
            column_f = sheet_obj_old_year['F']
            column_aj = sheet_obj_old_year['AJ']
            column_source_g = sheet_obj_sources_old_year['G']
            column_source_f = sheet_obj_sources_old_year['F']
            column_kids_d = sheet_obj_kids_old_year['C3'].value
            column_vmp_b = sheet_obj_vmp_old_year['B']

            calculating_excel(column_d, column_v, column_m, column_f, column_k, column_p, column_e, column_n, column_g,
                              column_aj, column_source_g, column_source_f, column_kids_d, column_vmp_b,
                              summ_1g, summ_1g_1, summ_1g_2, summ_2g, summ_2g_1, summ_2g_2, summ_3g, summ_4g, summ_4g_1,
                              summ_4g_2, summ_4g_3, summ_4g_4, summ_5g, summ_6g, summ_6g_1, summ_7g, summ_7g_1, summ_7g_2,
                              summ_7g_5, summ_7g_4, summ_8g, summ_8g_1, summ_8g_2, summ_8g_3, summ_8g_4, summ_9g, summ_10g,
                              summ_11g, summ_11g_1, summ_11g_2, summ_11g_3, summ_11g_4, summ_11g_5, summ_11g_6, summ_11g_7,
                              summ_11g_8, summ_11g_9, summ_12g, summ_12g_1, summ_vmp, summ_oms, summ_pd, uniq_pers_deti,
                              summ_aproba, count_sources, total_konserv, summ_foreign, summ_vmp_oms_from_begin_year,
                              summ_pleopto_kids_int_c006, summ_13g, summ_1g_previous_m_year)




def prev_year(column_d, column_v, column_m, column_f, column_k, column_p, column_e, column_n, column_g,
                      column_aj, column_source_g, column_source_f, column_kids_d, column_vmp_b,
                      summ_1g, summ_1g_1, summ_1g_2, summ_2g, summ_2g_1, summ_2g_2, summ_3g, summ_4g, summ_4g_1,
                      summ_4g_2, summ_4g_3, summ_4g_4, summ_5g, summ_6g, summ_6g_1, summ_7g, summ_7g_1, summ_7g_2,
                      summ_7g_5, summ_7g_4, summ_8g, summ_8g_1, summ_8g_2, summ_8g_3, summ_8g_4, summ_9g, summ_10g,
                      summ_11g, summ_11g_1, summ_11g_2, summ_11g_3, summ_11g_4, summ_11g_5, summ_11g_6, summ_11g_7,
                      summ_11g_8, summ_11g_9, summ_12g, summ_12g_1, summ_vmp, summ_oms, summ_pd, uniq_pers_deti,
                      summ_aproba, count_sources, total_konserv, summ_foreign, summ_vmp_oms_from_begin_year,
                      summ_pleopto_kids_int_c006, summ_13g, set_uniq_pers, set_uniq_pers_kons_treatment,
                      set_list_foreign_peoples, summ_1g_previous_m_year=0):
    pass



def calculating_excel(column_d, column_v, column_m, column_f, column_k, column_p, column_e, column_n, column_g,
                      column_aj, column_source_g, column_source_f, column_kids_d, column_vmp_b,
                      summ_1g, summ_1g_1, summ_1g_2, summ_2g, summ_2g_1, summ_2g_2, summ_3g, summ_4g, summ_4g_1,
                      summ_4g_2, summ_4g_3, summ_4g_4, summ_5g, summ_6g, summ_6g_1, summ_7g, summ_7g_1, summ_7g_2,
                      summ_7g_5, summ_7g_4, summ_8g, summ_8g_1, summ_8g_2, summ_8g_3, summ_8g_4, summ_9g, summ_10g,
                      summ_11g, summ_11g_1, summ_11g_2, summ_11g_3, summ_11g_4, summ_11g_5, summ_11g_6, summ_11g_7,
                      summ_11g_8, summ_11g_9, summ_12g, summ_12g_1, summ_vmp, summ_oms, summ_pd, uniq_pers_deti,
                      summ_aproba, count_sources, total_konserv, summ_foreign, summ_vmp_oms_from_begin_year,
                      summ_pleopto_kids_int_c006, summ_13g, summ_1g_previous_m_year=0):
    # расчёты:
    set_uniq_pers = set()
    set_uniq_pers_kons_treatment = set()
    set_list_foreign_peoples = set()

    summ_1g_2_int = two_in_one_cell(column_d, column_v, list_for_nums_hk, list_for_nums_iol_1,
                                    'Хирургия катаракты',
                                    'Инжекторная имплантация ИОЛ через разрез 1.8 мм',
                                    'Инжекторная имплантация ИОЛ через разрез 2 мм')

    summ_7g_2_int = two_in_one_cell(column_d, column_v, list_for_nums_end, list_for_nums_25g_1,
                                    'Эндовитреальная хирургия', 'Полная хирургия через системы 25 G')

    summ_7g_3_int = two_in_one_cell(column_d, column_v, list_for_nums_end_2, list_for_nums_27g_1,
                                    'Эндовитреальная хирургия', 'Полная хирургия через системы 27 G')

    summ_11g_4_int = two_in_one_cell(column_d, column_v, list_for_nums_vhk, list_for_nums_25g_2,
                                    'Витреоретинальная хирургия+хирургия катаракты',
                                    'Полная хирургия через системы 25 G')

    summ_11g_5_int = two_in_one_cell(column_d, column_v, list_for_nums_vhhk, list_for_nums_27g_2,
                                    'Витреоретинальная хирургия+хирургия катаракты',
                                    'Полная хирургия через системы 27 G')

    for i, row in enumerate(column_d):
        if row.value == 'Курсы консервативного лечения':
            list_for_nums_of_rows_konserv.append(i)
    for i, row in enumerate(column_m):
        if i == 0:
            continue
        if int(row.value) < 18:
            list_for_nums_of_rows_deti.append(i)
    summ_12g_1_int = len(intersection_list(list_for_nums_of_rows_konserv, list_for_nums_of_rows_deti))
    summ_11g_3_int = summ_11g_5_int + summ_11g_4_int

    for i, row in enumerate(column_f):
        try:
            if row.value.startswith('Курс плеоптики') or row.value.startswith('Курс плеопто-ортоптического лечения'):
                list_for_nums_pleoptika.append(i)
        except AttributeError:
            print(i, ' ', row.value)
    summ_pleopto_kids_int = len(intersection_list(list_for_nums_pleoptika, list_for_nums_of_rows_deti))

    for i, row in enumerate(column_k):
        if row.value.startswith('ВМП - медицинские услуги, оказанные за счет средств федерального'):
            summ_vmp += 1

    for i, row in enumerate(column_k):
        if row.value.startswith('ОМС') or row.value.startswith('СМП') or row.value == \
                'ВМП - медицинские услуги, оказанные за счет средств ФФОМС':
            summ_oms += 1

    for i, row in enumerate(column_k):
        if row.value.startswith('ПД') or row.value == 'ДМС' or row.value == 'Прочее':
            summ_pd += 1
    for i, row in enumerate(column_k):
        if row.value.startswith('Медицинские услуги'):
            summ_aproba += 1

    for i, row in enumerate(column_p):
        if row.value not in set_uniq_pers:
            set_uniq_pers.add(row.value)

    for i, row in enumerate(column_d):
        if row.value.startswith('Курсы консервативного лечения'):
            list_for_nums_of_rows_kk.append(i)
    for i, row in enumerate(column_p):
        if i in list_for_nums_of_rows_kk:
            set_uniq_pers_kons_treatment.add(row.value)

    for i, row in enumerate(column_k):
        if row.value.startswith('ВМП - медицинские услуги, оказанные за счет средств федерального'):
            list_for_nums_of_rows_vmp.append(i)
    for i, row in enumerate(column_n):
        if row.value == ' Дети':
            list_for_nums_of_rows_deti_vmp.append(i)
    summ_14g_3_int = len(intersection_list(list_for_nums_of_rows_vmp, list_for_nums_of_rows_deti_vmp))

    for i, row in enumerate(column_p):
        if i in list_for_nums_of_rows_deti_vmp:
            if row.value in set_uniq_pers:
                all_card_kids.append(row.value)
                if i not in list_for_nums_of_rows_kk:
                    list_for_nums_oper_deti.append(i)
    uniq_pers_deti = len(set(all_card_kids))

    list_for_nums_of_operated_patients = []
    # RUN IN SOURCES
    for i, row in enumerate(column_source_g):
        if i == 0:
            continue
        count_sources += int(row.value)

    for i, row in enumerate(column_source_f):
        if row.value == 'Конcерв':
            list_for_nums_of_operated_patients.append(i)
    for i, row in enumerate(column_source_g):
        if i in list_for_nums_of_operated_patients:
            total_konserv += int(row.value)

    for i, row in enumerate(column_aj):
        if row.value == 1:
            list_foreign_peoples.append(i)
    for i, row in enumerate(column_p):
        if i in list_foreign_peoples:
            set_list_foreign_peoples.add(row.value)
    summ_foreign = len(set_list_foreign_peoples)

    for i, row in enumerate(column_vmp_b):
        if row.value == 'II':
            summ_vmp_oms_from_begin_year += 1


    for i, row in enumerate(column_e):

        # print('***', row.value, '***') # RESULT > *** 24 ***
        if row.value in ('A018.04', 'A001.43', 'A018.03', 'A015.01', 'A001.41', 'A018.06', 'A001.76', 'A018.08',
                         'A001.84', 'A018.12', 'A001.116', 'A001.95', 'A018.10', 'A001.29', 'A001.13', 'A001.96',
                         'A001.65', 'A001.67', 'A001.32', 'A001.71', 'A001.79', 'A001.112', 'A001.90', 'A001.120',
                         'A001.42', 'A001.40', 'A001.75', 'A001.83', 'A001.115', 'A001.94', 'A001.27', 'A001.36',
                         'A001.28', 'A001.34', 'A001.74', 'A001.82', 'A001.114', 'A001.93', 'A001.25', 'A001.37',
                         'A001.30', 'A001.33', 'A001.69', 'A001.77', 'A001.88', 'A001.119', 'A001.26', 'A001.31',
                         'A001.132'):
            summ_1g += 1
        # если ежемесячный отчёт:
        if request.method == 'POST':
            if request.form['submit'] == "Ежемесячный":
                #summ_1g_previous_m_year =
                pass


        if row.value in ('A018.04', 'A001.43', 'A018.03', 'A015.01', 'A001.41', 'A018.06', 'A001.76', 'A018.08',
                         'A001.84', 'A018.12', 'A001.116', 'A001.95', 'A018.10', 'A001.29', 'A001.13', 'A001.96',
                         'A001.65', 'A001.67', 'A001.32', 'A001.71', 'A001.79', 'A001.112', 'A001.90', 'A001.120',
                         'A001.42'):
            summ_1g_1 += 1
        if row.value in ('A003.36', 'A003.07', 'A003.04', 'A003.03', 'A003.05', 'A003.13', 'A003.12', 'A008.11',
                         'A003.30', 'A003.21', 'A003.16', 'A003.24', 'A003.01', 'A003.22', 'A003.19', 'A003.29',
                         'A003.26', 'A003.31', 'A003.18', 'A015.02', 'A003.25', 'A008.09', 'A001.54'):
            summ_2g += 1
        if row.value in ('A003.36', 'A003.07', 'A003.04', 'A003.03', 'A003.05', 'A003.13', 'A003.12'):
            summ_2g_1 += 1
        if row.value in ('A008.11', 'A003.30', 'A003.21', 'A003.16', 'A003.24', 'A003.01', 'A003.22', 'A003.19',
                         'A003.29', 'A003.26', 'A003.31', 'A003.18', 'A015.02', 'A003.25', 'A008.09', 'A001.54'):
            summ_2g_2 += 1
        if row.value in ('A008.06', 'A008.05', 'AC008.01', 'AC008.02'):
            summ_3g += 1
        if row.value in ('B006.11', 'B006.03', 'BO06.01', 'B006.05', 'B006.02', 'B006.04', 'B006.32', 'B006.18',
                         'B006.23', 'B006.27', 'B006.16', 'B006.31', 'B006.30', 'B006.26', 'B006.29', 'B006.28',
                         'B006.20', 'B006.17', 'B006.19', 'B006.22', 'B006.21', 'B006.12', 'B006.06', 'B006.10',
                         'B006.09', 'B006.14', 'B006.08', 'B006.07', 'B006.15'):
            summ_4g += 1
        if row.value in ('B006.11', 'B006.03', 'BO06.01', 'B006.05', 'B006.02', 'B006.04', 'B006.32', 'B006.18',
                         'B006.23', 'B006.27', 'B006.16', 'B006.31', 'B006.30', 'B006.26', 'B006.29', 'B006.28',
                         'B006.20', 'B006.17', 'B006.19', 'B006.22', 'B006.21'):
            summ_4g_1 += 1
        if row.value in ('B006.32', 'B006.18', 'B006.23', 'B006.27', 'B006.16', 'B006.31', 'B006.30', 'B006.26',
                         'B006.29', 'B006.28', 'B006.20', 'B006.17', 'B006.19'):
            summ_4g_2 += 1
        if row.value in ('B006.12', 'B006.06', 'B006.10'):
            summ_4g_3 += 1
        if row.value in ('B006.09', 'B006.14', 'B006.08', 'B006.07', 'B006.15'):
            summ_4g_4 += 1
        if row.value in ('A001.23', 'A001.86', 'A001.22', 'A001.85', 'A006.02', 'A006.01', 'A006.04', 'A006.03',
                         'A006.05', 'A001.64', 'A001.14', 'A001.66', 'A001.68', 'A001.38', 'A001.70', 'A001.78',
                         'A001.110', 'A001.89'):
            summ_5g += 1
        if row.value in ('A001.107', 'A001.48', 'A001.10', 'A001.11', 'A001.08', 'A001.118', 'A001.127', 'A001.02',
                         'A001.01', 'A001.108', 'B006.13', 'A001.129', 'A001.234', 'A001.128', 'A001.233', 'A001.126',
                         'A001.12', 'A001.131', 'A001.09', 'A001.60'):
            summ_6g += 1
        if row.value in ('A001.107', 'A001.48', 'A001.10', 'A001.11', 'A001.08', 'A001.127', 'A001.02', 'A001.01',
                         'A001.108', 'B006.13', 'A001.129', 'A001.234', 'A001.128', 'A001.233', 'A001.126', 'A001.12',
                         'A001.09', 'A001.60'):
            summ_6g_1 += 1
        if row.value in ('A002.26', 'A002.44', 'A002.37', 'A002.38', 'A002.01', 'A002.25', 'A002.27', 'A002.03',
                         'A002.33', 'A002.32', 'A002.31', 'A002.34', 'A002.05', 'A002.14', 'A002.10', 'A002.40',
                         'A002.12', 'A002.13', 'A002.11', 'A002.39', 'A002.04', 'A002.08', 'A002.02', 'A002.47'):
            summ_7g_1 += 1
        if row.value in ('A002.26', 'A002.44', 'A002.37', 'A002.38', 'A002.01', 'A002.25', 'A002.27', 'A002.03',
                         'A002.33', 'A002.32', 'A002.31', 'A002.34', 'A002.05', 'A002.14', 'A002.10', 'A002.40',
                         'A002.12', 'A002.13', 'A002.11', 'A002.39', 'A002.04', 'A002.08', 'A002.02', 'A002.35',
                         'A002.17', 'A002.23', 'A002.18', 'A002.24', 'A002.20', 'A002.30', 'A002.21', 'A002.29',
                         'A002.28', 'A002.36', 'A002.45', 'A002.22', 'A002.19', 'A002.47'):
            summ_7g += 1
        if row.value in ('A002.35', 'A002.17', 'A002.23', 'A002.18', 'A002.24', 'A002.20', 'A002.30', 'A002.21',
                         'A002.29', 'A002.28', 'A002.22', 'A002.19'):
            summ_7g_4 += 1
        if row.value in ('A002.36', 'A002.45'):
            summ_7g_5 += 1
        if row.value in ('B002.08', 'B002.16', 'B002.02', 'B002.18', 'B002.04', 'B002.12', 'B002.06', 'B002.03',
                         'B002.07', 'B002.05', 'B002.25', 'B002.24', 'B002.21', 'B002.23', 'B002.22', 'B002.09',
                         'B002.10', 'B001.01', 'B001.03', 'B001.17', 'B001.07', 'B002.14', 'B001.06', 'B002.15',
                         'B001.11', 'B001.05', 'B003.04', 'B003.09', 'B003.07', 'B003.10', 'B003.05', 'B003.06',
                         'B003.08', 'B001.04', 'B001.02', 'B003.02', 'B003.03', 'B003.11', 'B003.01', 'B002.11',
                         'B001.08', 'B004.03', 'B001.14', 'B004.01', 'B002.17', 'B001.15', 'B001.09', 'B001.13',
                         'B002.19', 'B004.04', 'B002.20', 'B004.02', 'B009.07', 'B009.05', 'B009.02', 'B009.06',
                         'B009.03', 'B009.04', 'B002.01', 'B001.16', 'B001.10', 'B001.12'):
            summ_8g += 1
        if row.value in ('B002.08', 'B002.16', 'B002.02', 'B002.18', 'B002.04', 'B002.12', 'B002.06', 'B002.03',
                         'B002.07', 'B002.05', 'B002.25', 'B002.24', 'B002.21', 'B002.23', 'B002.22', 'B002.09',
                         'B002.10'):
            summ_8g_1 += 1
        if row.value in ('B001.01', 'B001.03', 'B001.17', 'B001.07', 'B002.14', 'B001.06', 'B002.15', 'B001.11',
                         'B001.05'):
            summ_8g_2 += 1
        if row.value in ('B003.04', 'B003.09', 'B003.07', 'B003.10', 'B003.05', 'B003.06', 'B003.08', 'B001.04',
                         'B001.02', 'B003.02', 'B003.03', 'B003.11', 'B003.01'):
            summ_8g_3 += 1
        if row.value in ('B002.11', 'B001.08', 'B004.03', 'B001.14', 'B004.01', 'B002.17', 'B001.15', 'B001.09',
                         'B001.13', 'B002.19', 'B004.04', 'B002.20', 'B004.02', 'B009.07', 'B009.05', 'B009.02',
                         'B009.06', 'B009.03', 'B009.04', 'B002.01', 'B001.16', 'B001.10', 'B001.12'):
            summ_8g_4 += 1
        if row.value in ('AC005.01', 'AC005.02', 'A005.36', 'AC005.03', 'A005.39', 'A005.27', 'A005.29', 'A005.28',
                         'A005.19', 'A005.59', 'A005.78', 'A005.08', 'A005.61', 'A005.47', 'A005.56', 'A005.18',
                         'A005.34', 'A005.62', 'A005.03', 'A005.55', 'A005.13', 'A005.75', 'A005.58', 'A005.76',
                         'A005.17', 'A005.52', 'A005.12', 'A005.51', 'A005.25', 'A005.05', 'A005.53', 'A005.30',
                         'A005.15', 'A005.26', 'A005.60', 'A005.46', 'A005.07', 'A005.06', 'A005.01', 'A005.79',
                         'A005.45', 'A005.69', 'A005.23', 'AF005.22', 'AC005.04', 'A005.24', 'A005.33', 'A005.50',
                         'A005.14', 'A005.68', 'A005.32', 'A005.63', 'A005.71', 'A005.64', 'A005.42', 'ACO005.05',
                         'A005.41', 'A005.16', 'A005.38', 'C005.06', 'A005.40', 'A005.72', 'A005.37', 'A005.73',
                         'A005.57', 'A005.77', 'A005.31', 'A005.54', 'A005.49', 'A005.44', 'A005.04', 'A005.70',
                         'A005.11', 'A005.09', 'A005.10', 'A005.20', 'A005.43', 'A005.21', 'A005.35'):
            summ_9g += 1
        if row.value in ('AC004.06', 'A004.22', 'A004.19', 'A004.05', 'A004.04', 'A004.07', 'A004.31', 'A004.29',
                         'A004.30', 'A004.21', 'A004.25', 'A004.09', 'A004.06', 'ABO004.01', 'AC004.01', 'AC004.03',
                         'AC004.04', 'AC004.02', 'A004.13', 'A004.02', 'A004.10', 'A004.28', 'A004.24', 'A004.23',
                         'A004.20', 'A004.15', 'A004.18', 'A004.14', 'AC004.05', 'A004.34', 'A004.12', 'A004.27',
                         'A004.26', 'A004.35', 'A004.32', 'A004.33', 'A004.01', 'A004.11', 'A004.03', 'A004.16',
                         'A004.17'):
            summ_10g += 1
        if row.value in ('A012.05', 'A012.06', 'A012.66', 'A012.07', 'A012.08', 'A012.23', 'A012.26', 'A012.79',
                         'A012.29', 'A013.27', 'A012.70', 'A012.72', 'A012.74', 'A012.76', 'A012.78', 'A012.19',
                         'A012.35', 'A012.36', 'A012.37', 'A012.68', 'A012.69', 'A012.71', 'A012.73', 'A012.75',
                         'A012.77', 'A012.38', 'A012.39', 'A012.40', 'A012.41', 'A012.56', 'A012.58', 'A012.60',
                         'A012.82', 'A012.62', 'A012.16', 'A012.12', 'A002.42', 'A002.43', 'A123.12', 'A123.13',
                         'A012.09', 'A012.24', 'A012.27', 'A012.10', 'A012.11', 'A002.41', 'A012.65', 'A012.64',
                         'A123.14', 'A123.15', 'A012.13', 'A123.07', 'A012.14', 'A123.08', 'A012.25', 'A123.09',
                         'A012.28', 'A123.16', 'A012.80', 'A123.10', 'A012.31', 'A123.11', 'A012.15', 'A012.54',
                         'A012.55', 'A012.57', 'A012.59', 'A012.61', 'A012.83', 'A012.63', 'A012.67', 'A012.48',
                         'A012.50', 'A012.51', 'A012.81', 'A012.53', 'A012.52', 'A013.32', 'A013.35', 'A013.28',
                         'A013.29', 'A013.30', 'A013.31', 'A013.34', 'A013.07', 'A013.11', 'A013.13', 'A013.16',
                         'A013.33', 'A013.05', 'A013.12', 'A013.14', 'A013.26', 'A013.15', 'A013.06', 'A013.04',
                         'A001.245', 'A001.246', 'A001.247', 'A001.242', 'A001.243', 'A001.244', 'A001.239', 'A001.240',
                         'A001.241', 'A001.237', 'A001.238', 'A001.236', 'A001.97', 'A001.102', 'A001.103', 'A001.104',
                         'A001.105', 'A001.03', 'A001.98', 'A001.99', 'A001.100', 'A001.101', 'A001.04', 'A001.61',
                         'A001.62', 'A001.63'):
            summ_11g += 1
        if row.value in ('A012.05', 'A012.06', 'A012.66', 'A012.07', 'A012.08', 'A012.23', 'A012.26', 'A012.79',
                         'A012.29', 'A013.27', 'A012.70', 'A012.72', 'A012.74', 'A012.76', 'A012.78', 'A012.19',
                         'A012.35', 'A012.36', 'A012.37', 'A012.68', 'A012.69', 'A012.71', 'A012.73', 'A012.75',
                         'A012.77', 'A012.38', 'A012.39', 'A012.40', 'A012.41', 'A012.56', 'A012.58', 'A012.60',
                         'A012.82', 'A012.62', 'A012.16', 'A012.12', 'A002.42', 'A002.43', 'A123.12', 'A123.13',
                         'A012.09', 'A012.24', 'A012.27', 'A012.10', 'A012.11', 'A002.41', 'A012.65', 'A012.64',
                         'A123.14', 'A123.15', 'A012.13', 'A123.07', 'A012.14', 'A123.08', 'A012.25', 'A123.09',
                         'A012.28', 'A123.16', 'A012.80', 'A123.10', 'A012.31', 'A123.11', 'A012.15', 'A012.54',
                         'A012.55', 'A012.57', 'A012.59', 'A012.61', 'A012.83', 'A012.63', 'A012.67', 'A012.48',
                         'A012.50', 'A012.51', 'A012.81', 'A012.53', 'A012.52', 'A013.05'):
            summ_11g_1 += 1
        if row.value in ('A012.05', 'A012.06', 'A012.66', 'A012.07', 'A012.08', 'A012.23', 'A012.26', 'A012.79',
                         'A012.29', 'A013.27', 'A012.70', 'A012.72', 'A012.74', 'A012.76', 'A012.78', 'A012.19',
                         'A012.35', 'A012.36', 'A012.37', 'A012.68', 'A012.69', 'A012.71', 'A012.73', 'A012.75',
                         'A012.77', 'A012.38', 'A012.39', 'A012.40', 'A012.41', 'A012.56', 'A012.58', 'A012.60',
                         'A012.82', 'A012.62', 'A012.16', 'A012.12', 'A002.42', 'A002.43', 'A123.12', 'A123.13',
                         'A012.09', 'A012.24', 'A012.27', 'A012.10', 'A012.11', 'A002.41', 'A012.65', 'A012.64',
                         'A123.14', 'A123.15', 'A012.13', 'A123.07', 'A012.14', 'A123.08', 'A012.25', 'A123.09',
                         'A012.28', 'A123.16', 'A012.80', 'A123.10', 'A012.31', 'A123.11', 'A012.15', 'A012.54',
                         'A012.55', 'A012.57', 'A012.59', 'A012.61', 'A012.83', 'A012.63', 'A012.67', 'A012.48',
                         'A012.50', 'A012.51', 'A012.81', 'A012.53', 'A012.52', 'A013.05'):
            summ_11g_2 += 1
        if row.value in ('A013.32', 'A013.35', 'A013.28', 'A013.29', 'A013.30', 'A013.31', 'A013.34', 'A013.07',
                         'A013.11', 'A013.13', 'A013.16', 'A013.33', 'A013.05', 'A013.12', 'A013.14', 'A013.26',
                         'A013.15', 'A013.06', 'A013.04'):
            summ_11g_6 += 1
        if row.value in ('A001.242', 'A001.243', 'A001.244', 'A001.237', 'A001.238', 'A001.236', 'A001.245', 'A001.246',
                         'A001.247', 'A001.242', 'A001.243', 'A001.244', 'A001.239', 'A001.240',
                         'A001.241', 'A001.237', 'A001.238', 'A001.236', 'A001.97', 'A001.102', 'A001.103', 'A001.104',
                         'A001.105', 'A001.03', 'A001.98', 'A001.99', 'A001.100', 'A001.101', 'A001.04', 'A001.61',
                         'A001.62', 'A001.63'):
            summ_11g_8 += 1
        #if row.value in ():
        #    summ_11g_9 += 1
        if row.value in ('C006.06', 'C006.11', 'C006.07', 'C006.01', 'C006.09', 'C006.04', 'C004.05', 'C004.04',
                         'C000.01', 'C000.02', 'C000.06', 'C000.07', 'C000.04', 'C001.16', 'C005.05', 'C005.03',
                         'C005.09', 'C002.24', 'C003.02', 'C003.01', 'C001.18', 'C001.04', 'C002.07', 'C004.09',
                         'C001.10', 'C002.36', 'C001.08', 'C001.05', 'C002.09', 'C002.04', 'C001.06', 'C012.06',
                         'C012.07', 'C001.14', 'C002.11', 'C001.11', 'C002.16', 'C001.17', 'C004.03', 'C004.02',
                         'C002.37', 'C012.03', 'C002.40', 'C004.06', 'C002.28', 'C002.38', 'C002.26', 'C002.06',
                         'C002.20', 'C002.18', 'C001.02', 'C001.01', 'C012.09', 'C005.07', 'C002.08', 'C002.03',
                         'C002.10', 'C012.05', 'C001.09', 'C001.13', 'C006.10', 'C006.05', 'C006.12', 'C002.34',
                         'C002.23', 'C002.35', 'C001.03', 'C009.05', 'C009.06', 'C009.04', 'C009.03', 'C009.07'):
            summ_12g += 1

        if row.value not in ('A018.04', 'A001.43', 'A018.03', 'A015.01', 'A001.41', 'A018.06', 'A001.76', 'A018.08',
                         'A001.84', 'A018.12', 'A001.116', 'A001.95', 'A018.10', 'A001.29', 'A001.13', 'A001.96',
                         'A001.65', 'A001.67', 'A001.32', 'A001.71', 'A001.79', 'A001.112', 'A001.90', 'A001.120',
                         'A001.42', 'A001.40', 'A001.75', 'A001.83', 'A001.115', 'A001.94', 'A001.27', 'A001.36',
                         'A001.28', 'A001.34', 'A001.74', 'A001.82', 'A001.114', 'A001.93', 'A001.25', 'A001.37',
                         'A001.30', 'A001.33', 'A001.69', 'A001.77', 'A001.88', 'A001.119', 'A001.26', 'A001.31',
                         'A001.132','A003.36', 'A003.07', 'A003.04', 'A003.03', 'A003.05', 'A003.13', 'A003.12',
                        'A008.11',
                         'A003.30', 'A003.21', 'A003.16', 'A003.24', 'A003.01', 'A003.22', 'A003.19', 'A003.29',
                         'A003.26', 'A003.31', 'A003.18', 'A015.02', 'A003.25', 'A008.06', 'A008.05', 'AC008.01',
                        'AC008.02',
                        'B006.11', 'B006.03', 'BO06.01', 'B006.05', 'B006.02', 'B006.04', 'B006.32', 'B006.18',
                         'B006.23', 'B006.27', 'B006.16', 'B006.31', 'B006.30', 'B006.26', 'B006.29', 'B006.28',
                         'B006.20', 'B006.17', 'B006.19', 'B006.22', 'B006.21', 'B006.12', 'B006.06', 'B006.10',
                         'B006.09', 'B006.14', 'B006.08', 'B006.07', 'B006.15',
                        'A001.23', 'A001.86', 'A001.22', 'A001.85', 'A006.02', 'A006.01', 'A006.04', 'A006.03',
                         'A006.05', 'A001.64', 'A001.14', 'A001.66', 'A001.68', 'A001.38', 'A001.70', 'A001.78',
                         'A001.110', 'A001.89','A001.107', 'A001.48', 'A001.10', 'A001.11', 'A001.08', 'A001.118',
                        'A001.127', 'A001.02',
                         'A001.01', 'A001.108', 'B006.13', 'A001.129', 'A001.234', 'A001.128', 'A001.233', 'A001.126',
                         'A001.12', 'A001.131', 'A002.26', 'A002.44', 'A002.37', 'A002.38', 'A002.01', 'A002.25',
                        'A002.27', 'A002.03',
                         'A002.33', 'A002.32', 'A002.31', 'A002.34', 'A002.05', 'A002.14', 'A002.10', 'A002.40',
                         'A002.12', 'A002.13', 'A002.11', 'A002.39', 'A002.04', 'A002.08', 'A002.02', 'A002.35',
                         'A002.17', 'A002.23', 'A002.18', 'A002.24', 'A002.20', 'A002.30', 'A002.21', 'A002.29',
                         'A002.28', 'A002.36', 'A002.45', 'A002.22', 'A002.19', 'B002.08', 'B002.16', 'B002.02',
                        'B002.18', 'B002.04', 'B002.12', 'B002.06', 'B002.03',
                         'B002.07', 'B002.05', 'B002.25', 'B002.24', 'B002.21', 'B002.23', 'B002.22', 'B002.09',
                         'B002.10', 'B001.01', 'B001.03', 'B001.17', 'B001.07', 'B002.14', 'B001.06', 'B002.15',
                         'B001.11', 'B001.05', 'B003.04', 'B003.09', 'B003.07', 'B003.10', 'B003.05', 'B003.06',
                         'B003.08', 'B001.04', 'B001.02', 'B003.02', 'B003.03', 'B003.11', 'B003.01', 'B002.11',
                         'B001.08', 'B004.03', 'B001.14', 'B004.01', 'B002.17', 'B001.15', 'B001.09', 'B001.13',
                         'B002.19', 'B004.04', 'B002.20', 'B004.02', 'B009.07', 'B009.05', 'B009.02', 'B009.06',
                         'B009.03', 'B009.04', 'B002.01', 'B001.16', 'B001.10', 'B001.12', 'AC005.01', 'AC005.02',
                        'A005.36', 'AC005.03', 'A005.39', 'A005.27', 'A005.29', 'A005.28',
                         'A005.19', 'A005.59', 'A005.78', 'A005.08', 'A005.61', 'A005.47', 'A005.56', 'A005.18',
                         'A005.34', 'A005.62', 'A005.03', 'A005.55', 'A005.13', 'A005.75', 'A005.58', 'A005.76',
                         'A005.17', 'A005.52', 'A005.12', 'A005.51', 'A005.25', 'A005.05', 'A005.53', 'A005.30',
                         'A005.15', 'A005.26', 'A005.60', 'A005.46', 'A005.07', 'A005.06', 'A005.01', 'A005.79',
                         'A005.45', 'A005.69', 'A005.23', 'AF005.22', 'AC005.04', 'A005.24', 'A005.33', 'A005.50',
                         'A005.14', 'A005.68', 'A005.32', 'A005.63', 'A005.71', 'A005.64', 'A005.42', 'ACO005.05',
                         'A005.41', 'A005.16', 'A005.38', 'C005.06', 'A005.40', 'A005.72', 'A005.37', 'A005.73',
                         'A005.57', 'A005.77', 'A005.31', 'A005.54', 'A005.49', 'A005.44', 'A005.04', 'A005.70',
                         'A005.11', 'A005.09', 'A005.10', 'A005.20', 'A005.43', 'A005.21', 'A005.35', 'AC004.06',
                        'A004.22', 'A004.19', 'A004.05', 'A004.04', 'A004.07', 'A004.31', 'A004.29',
                         'A004.30', 'A004.21', 'A004.25', 'A004.09', 'A004.06', 'ABO004.01', 'AC004.01', 'AC004.03',
                         'AC004.04', 'AC004.02', 'A004.13', 'A004.02', 'A004.10', 'A004.28', 'A004.24', 'A004.23',
                         'A004.20', 'A004.15', 'A004.18', 'A004.14', 'AC004.05', 'A004.34', 'A004.12', 'A004.27',
                         'A004.26', 'A004.35', 'A004.32', 'A004.33', 'A004.01', 'A004.11', 'A004.03', 'A004.16',
                         'A004.17', 'A012.05', 'A012.06', 'A012.66', 'A012.07', 'A012.08', 'A012.23', 'A012.26',
                        'A012.79',
                         'A012.29', 'A013.27', 'A012.70', 'A012.72', 'A012.74', 'A012.76', 'A012.78', 'A012.19',
                         'A012.35', 'A012.36', 'A012.37', 'A012.68', 'A012.69', 'A012.71', 'A012.73', 'A012.75',
                         'A012.77', 'A012.38', 'A012.39', 'A012.40', 'A012.41', 'A012.56', 'A012.58', 'A012.60',
                         'A012.82', 'A012.62', 'A012.16', 'A012.12', 'A002.42', 'A002.43', 'A123.12', 'A123.13',
                         'A012.09', 'A012.24', 'A012.27', 'A012.10', 'A012.11', 'A002.41', 'A012.65', 'A012.64',
                         'A123.14', 'A123.15', 'A012.13', 'A123.07', 'A012.14', 'A123.08', 'A012.25', 'A123.09',
                         'A012.28', 'A123.16', 'A012.80', 'A123.10', 'A012.31', 'A123.11', 'A012.15', 'A012.54',
                         'A012.55', 'A012.57', 'A012.59', 'A012.61', 'A012.83', 'A012.63', 'A012.67', 'A012.48',
                         'A012.50', 'A012.51', 'A012.81', 'A012.53', 'A012.52', 'A013.32', 'A013.35', 'A013.28',
                         'A013.29', 'A013.30', 'A013.31', 'A013.34', 'A013.07', 'A013.11', 'A013.13', 'A013.16',
                         'A013.33', 'A013.05', 'A013.12', 'A013.14', 'A013.26', 'A013.15', 'A013.06', 'A013.04',
                         'A001.245', 'A001.246', 'A001.247', 'A001.242', 'A001.243', 'A001.244', 'A001.239', 'A001.240',
                         'A001.241', 'A001.237', 'A001.238', 'A001.236', 'A001.97', 'A001.102', 'A001.103', 'A001.104',
                         'A001.105', 'A001.03', 'A001.98', 'A001.99', 'A001.100', 'A001.101', 'A001.04', 'A001.61',
                         'A001.62', 'A001.63', 'C006.06', 'C006.11', 'C006.07', 'C006.01', 'C006.09', 'C006.04',
                        'C004.05', 'C004.04',
                         'C000.01', 'C000.02', 'C000.06', 'C000.07', 'C000.04', 'C001.16', 'C005.05', 'C005.03',
                         'C005.09', 'C002.24', 'C003.02', 'C003.01', 'C001.18', 'C001.04', 'C002.07', 'C004.09',
                         'C001.10', 'C002.36', 'C001.08', 'C001.05', 'C002.09', 'C002.04', 'C001.06', 'C012.06',
                         'C012.07', 'C001.14', 'C002.11', 'C001.11', 'C002.16', 'C001.17', 'C004.03', 'C004.02',
                         'C002.37', 'C012.03', 'C002.40', 'C004.06', 'C002.28', 'C002.38', 'C002.26', 'C002.06',
                         'C002.20', 'C002.18', 'C001.02', 'C001.01', 'C012.09', 'C005.07', 'C002.08', 'C002.03',
                         'C002.10', 'C012.05', 'C001.09', 'C001.13', 'C006.10', 'C006.05', 'C006.12', 'C002.34',
                         'C002.23', 'C002.35', 'C001.03', 'C009.05', 'C009.06', 'C009.04', 'C009.03', 'C009.07',
                             'A001.09', 'A001.60', 'A008.09', 'A001.54', 'A002.47', 'ОперМск'):

            #if row.value in ('A001.46', 'A001.49'):  # 'A001.51', 'A001.81', 'A001.09', 'A001.49'
            summ_13g += 1
            message = f"Не в нашем перечне кодов: {row.value}"
            flash(message)
            print('Не в нашем перечне кодов: ', row.value)

        if row.value == 'C006.06':
            list_for_nums_of_rows_c006_kids.append(i)
        summ_pleopto_kids_int_c006 = len(intersection_list(list_for_nums_of_rows_c006_kids, list_for_nums_of_rows_deti_vmp))

        summa_all = summ_1g + summ_2g + summ_3g + summ_4g + summ_5g + summ_6g + summ_7g + summ_8g + summ_9g + \
                    summ_10g + summ_11g + summ_12g + summ_13g


        write_to_exel(summ_1g, summ_1g_1, summ_1g_2_int, summ_2g, summ_2g_1, summ_2g_2, summ_3g, summ_4g,
                      summ_4g_1, summ_4g_2, summ_4g_3, summ_4g_4, summ_5g, summ_6g, summ_6g_1, summ_7g,
                      summ_7g_1, summ_7g_2_int, summ_7g_3_int, summ_7g_4, summ_7g_5, summ_8g, summ_8g_1,
                      summ_8g_2, summ_8g_3, summ_8g_4, summ_9g, summ_10g, summ_11g, summ_11g_1, summ_11g_2,
                      summ_11g_3_int, summ_11g_4_int, summ_11g_5_int, summ_11g_6, summ_11g_8, summ_11g_9,
                      summ_12g, summ_12g_1_int, summ_13g, summa_all, count_sources, total_konserv, summ_foreign,
                      column_kids_d, list_for_nums_oper_deti, summ_pleopto_kids_int, summ_pleopto_kids_int_c006,
                      summ_pd, summ_vmp, summ_14g_3_int, summ_aproba, summ_oms, constant,
                      summ_vmp_oms_from_begin_year)


def write_to_exel(summ_1g, summ_1g_1, summ_1g_2_int, summ_2g, summ_2g_1, summ_2g_2, summ_3g, summ_4g,
                      summ_4g_1, summ_4g_2, summ_4g_3, summ_4g_4, summ_5g, summ_6g, summ_6g_1, summ_7g, summ_7g_1,
                      summ_7g_2_int, summ_7g_3_int, summ_7g_4, summ_7g_5, summ_8g, summ_8g_1, summ_8g_2, summ_8g_3,
                  summ_8g_4, summ_9g, summ_10g, summ_11g, summ_11g_1, summ_11g_2, summ_11g_3_int, summ_11g_4_int,
                  summ_11g_5_int, summ_11g_6, summ_11g_8, summ_11g_9, summ_12g, summ_12g_1_int,  summ_13g, summa_all,
                  count_sources, total_konserv, summ_foreign, column_kids_d, list_for_nums_oper_deti,
                  summ_pleopto_kids_int, summ_pleopto_kids_int_c006, summ_pd, summ_vmp, summ_14g_3_int, summ_aproba,
                  summ_oms, constant, summ_vmp_oms_from_begin_year, summ_1g_previous_m_year=0):

    if request.method == 'POST':
        if request.form['submit'] == "Сформировать еженедельный отчёт":
            sheet_obj_out['A2'].value = f'{date_from_day}.{date_from_month}.{date_from_year} - {date_to_day}.' \
                                        f'{date_to_month}.{date_to_year}'
            sheet_obj_out['B4'].value = summ_1g
            sheet_obj_out['B5'].value = summ_1g_1
            sheet_obj_out['B6'].value = summ_1g_2_int
            sheet_obj_out['B7'].value = summ_2g
            sheet_obj_out['B8'].value = summ_2g_1
            sheet_obj_out['B9'].value = summ_2g_2
            sheet_obj_out['B10'].value = summ_3g
            sheet_obj_out['B11'].value = summ_4g
            sheet_obj_out['B12'].value = summ_4g_1
            sheet_obj_out['B13'].value = summ_4g_2
            sheet_obj_out['B14'].value = summ_4g_3
            sheet_obj_out['B15'].value = summ_4g_4
            sheet_obj_out['B16'].value = summ_5g
            sheet_obj_out['B17'].value = summ_6g
            sheet_obj_out['B18'].value = summ_6g_1
            sheet_obj_out['B19'].value = summ_7g
            sheet_obj_out['B20'].value = summ_7g_1
            sheet_obj_out['B21'].value = summ_7g_2_int
            sheet_obj_out['B22'].value = summ_7g_3_int
            sheet_obj_out['B23'].value = summ_7g_4
            sheet_obj_out['B24'].value = summ_7g_5
            sheet_obj_out['B25'].value = summ_8g
            sheet_obj_out['B26'].value = summ_8g_1
            sheet_obj_out['B27'].value = summ_8g_2
            sheet_obj_out['B28'].value = summ_8g_3
            sheet_obj_out['B29'].value = summ_8g_4
            sheet_obj_out['B30'].value = summ_9g
            sheet_obj_out['B31'].value = summ_10g
            sheet_obj_out['B32'].value = summ_11g
            sheet_obj_out['B33'].value = summ_11g_1 + summ_11g_6
            sheet_obj_out['B34'].value = summ_11g_2
            sheet_obj_out['B35'].value = summ_11g_3_int
            sheet_obj_out['B36'].value = summ_11g_4_int
            sheet_obj_out['B37'].value = summ_11g_5_int
            sheet_obj_out['B38'].value = summ_11g_6
            sheet_obj_out['B39'].value = summ_11g_6
            sheet_obj_out['B40'].value = summ_11g_8
            sheet_obj_out['B41'].value = summ_11g_9
            sheet_obj_out['B42'].value = summ_12g
            sheet_obj_out['B43'].value = summ_12g_1_int
            sheet_obj_out['B44'].value = summ_13g
            sheet_obj_out['B46'].value = summa_all
            sheet_obj_out['A47'].value = 'Пролечено больных (всего)'
            sheet_obj_out['B47'].value = count_sources  # len(set_uniq_pers)
            sheet_obj_out['A48'].value = '    - прооперировано больных'
            sheet_obj_out['B48'].value = count_sources - total_konserv  # len(set_uniq_pers) - len(set_uniq_pers_kons_treatment)
            sheet_obj_out['A49'].value = '    - пролечено консервативно больных'
            sheet_obj_out['B49'].value = total_konserv  # len(set_uniq_pers_kons_treatment)
            sheet_obj_out['A50'].value = '    - пролечено иностранцев'
            sheet_obj_out['B50'].value = summ_foreign
            sheet_obj_out['A51'].value = 'количество обследованных пациентов до 18 лет'
            sheet_obj_out['B51'].value = column_kids_d
            sheet_obj_out['A52'].value = 'количество операций пациентам до 18 лет'
            sheet_obj_out['B52'].value = len(list_for_nums_oper_deti)
            sheet_obj_out['A53'].value = 'количество курсов плеопто-ортоптического лечения до 18 лет'
            sheet_obj_out['B53'].value = summ_pleopto_kids_int + summ_pleopto_kids_int_c006
            sheet_obj_out['A54'].value = '14. Операции и курсы лечения по видам финансирования:'
            sheet_obj_out['B54'].value = summa_all
            sheet_obj_out['A55'].value = '    - коммерческие'
            sheet_obj_out['B55'].value = summ_pd
            sheet_obj_out['A56'].value = '    - ВМП'
            sheet_obj_out['B56'].value = summ_vmp
            sheet_obj_out['A57'].value = 'из них дети'
            sheet_obj_out['B57'].value = summ_14g_3_int
            sheet_obj_out['A58'].value = '    - Клиническая апробация'
            sheet_obj_out['B58'].value = summ_aproba
            sheet_obj_out['A59'].value = '    - ОМС'
            sheet_obj_out['B59'].value = summ_oms
            sheet_obj_out['A60'].value = '15. Плановое задание бюджет (ВМП+Клиническая апробация) на 2022г.'
            sheet_obj_out['B60'].value = constant
            sheet_obj_out['A61'].value = '    - пролечено квот по бюджету (ВМП+Клиническая апробация) с начала года по наст. момент'
            sheet_obj_out['B61'].value = ''
            sheet_obj_out['A62'].value = '    - пролечено квот по ВМП с начала года по настоящий момент'
            sheet_obj_out['B62'].value = summ_vmp_oms_from_begin_year
            sheet_obj_out['A63'].value = '    - пролечено квот по Клиническая апробация с начала года по настоящий момент'
            sheet_obj_out['B63'].value = 0
            sheet_obj_out['A64'].value = ''
            sheet_obj_out['A65'].value = ''
        elif request.form['submit'] == "Ежемесячный":
            #reporting_m = reporting_month[date_from_month]
            sheet_obj_out['A2'].value = f'отчет по лечебной работе за {reporting_month[date_from_month]} месяц'
            sheet_obj_out['B3'].value = 2022
            sheet_obj_out['C3'].value = 2023
            sheet_obj_out['B4'].value = summ_1g_previous_m_year
            sheet_obj_out['C4'].value = summ_1g
            sheet_obj_out['C5'].value = summ_1g_1
            sheet_obj_out['C6'].value = summ_1g_2_int
            sheet_obj_out['C7'].value = summ_2g
            sheet_obj_out['C8'].value = summ_2g_1
            sheet_obj_out['C9'].value = summ_2g_2
            sheet_obj_out['C10'].value = summ_3g
            sheet_obj_out['C11'].value = summ_4g
            sheet_obj_out['C12'].value = summ_4g_1
            sheet_obj_out['C13'].value = summ_4g_2
            sheet_obj_out['C14'].value = summ_4g_3
            sheet_obj_out['C15'].value = summ_4g_4
            sheet_obj_out['C16'].value = summ_5g
            sheet_obj_out['C17'].value = summ_6g
            sheet_obj_out['C18'].value = summ_6g_1
            sheet_obj_out['C19'].value = summ_7g
            sheet_obj_out['C20'].value = summ_7g_1
            sheet_obj_out['C21'].value = summ_7g_2_int
            sheet_obj_out['C22'].value = summ_7g_3_int
            sheet_obj_out['C23'].value = summ_7g_4
            sheet_obj_out['C24'].value = summ_7g_5
            sheet_obj_out['C25'].value = summ_8g
            sheet_obj_out['C26'].value = summ_8g_1
            sheet_obj_out['C27'].value = summ_8g_2
            sheet_obj_out['C28'].value = summ_8g_3
            sheet_obj_out['C29'].value = summ_8g_4
            sheet_obj_out['C30'].value = summ_9g
            sheet_obj_out['C31'].value = summ_10g
            sheet_obj_out['C32'].value = summ_11g
            sheet_obj_out['C33'].value = summ_11g_1 + summ_11g_6
            sheet_obj_out['C34'].value = summ_11g_2
            sheet_obj_out['C35'].value = summ_11g_3_int
            sheet_obj_out['C36'].value = summ_11g_4_int
            sheet_obj_out['C37'].value = summ_11g_5_int
            sheet_obj_out['C38'].value = summ_11g_6
            sheet_obj_out['C39'].value = summ_11g_6
            sheet_obj_out['C40'].value = summ_11g_8
            sheet_obj_out['C41'].value = summ_11g_9
            sheet_obj_out['C42'].value = summ_12g
            sheet_obj_out['C43'].value = summ_12g_1_int
            sheet_obj_out['C44'].value = summ_13g
            sheet_obj_out['C46'].value = summa_all

        else:
            return 'Ошибка 023'

    sheet_obj_out['A4'].value = '1. Экстракция катаракты'
    sheet_obj_out['A5'].value = '    - факоэмульсификация'
    sheet_obj_out['A6'].value = '    - из них с размером 2,2 и менее'
    sheet_obj_out['A7'].value = '2. Антиглаукоматозные операции (всего):'
    sheet_obj_out['A8'].value = '    - микроинвазивные'
    sheet_obj_out['A9'].value = '    - др. хирургические АГО'
    sheet_obj_out['A10'].value = '3. Склероурепляющие операции'
    sheet_obj_out['A11'].value = '4. Лазерные рефракционные'
    sheet_obj_out['A12'].value = '    - Лазик'
    sheet_obj_out['A13'].value = '    в т.ч. Фемтолазик'
    sheet_obj_out['A14'].value = '    - ФРК'
    sheet_obj_out['A15'].value = '    - др. лазерные рефракционные'
    sheet_obj_out['A16'].value = '5. Нелазерные рефракционные'
    sheet_obj_out['A17'].value = '6. Операции на роговице'
    sheet_obj_out['A18'].value = '    - кератопластика с применением донорского материала'
    sheet_obj_out['A19'].value = '7. Витреоретинальные операции'
    sheet_obj_out['A20'].value = '    - из них эндовитреальные, в тч с витрэктомией'
    sheet_obj_out['A21'].value = '    - через системы 25 G'
    sheet_obj_out['A22'].value = '    - через систему 27 G'
    sheet_obj_out['A23'].value = '    - др. витреоретинальные'
    sheet_obj_out['A24'].value = '    - инъекции и/в'
    sheet_obj_out['A25'].value = '8. Лазерные нерефракционные'
    sheet_obj_out['A26'].value = '    - коагуляция сетчатки'
    sheet_obj_out['A27'].value = '    - ДЗК'
    sheet_obj_out['A28'].value = '    - по поводу глаукомы'
    sheet_obj_out['A29'].value = '    - др. лазерные нерефракционные'
    sheet_obj_out['A30'].value = '9. Окулопластические операции'
    sheet_obj_out['A31'].value = '10. Операции при опухолях'
    sheet_obj_out['A32'].value = '11. Сочетанные операции'
    sheet_obj_out['A33'].value = '    - сочетанные с ЭК'
    sheet_obj_out['A34'].value = '    - с разрезом 2,2 и менее'
    sheet_obj_out['A35'].value = '    - сочетанные с ЭВ'
    sheet_obj_out['A36'].value = '    - через системы 25-G'
    sheet_obj_out['A37'].value = '    - через системы 27-G'
    sheet_obj_out['A38'].value = '    - сочетанные с АГО'
    sheet_obj_out['A39'].value = '    - микро АГО'
    sheet_obj_out['A40'].value = '    - сочетанные с кератопластикой'
    sheet_obj_out['A41'].value = '    - другие сочетанные операции'
    sheet_obj_out['A42'].value = '12. Консервативное решение'
    sheet_obj_out['A43'].value = '    - дети'
    sheet_obj_out['A44'].value = '13. Прочие операции (всего)'
    sheet_obj_out['A45'].value = '    - из них прочие первичные'
    sheet_obj_out['B45'].value = '-'
    sheet_obj_out['A46'].value = 'ВСЕГО'



    # присваиваем итоговые значения, сохраняем в итоговый файл


def save_info_week(wb_obj_out):
    '''сохраняем в итоговый файл'''
    wb_obj_out.save("temp_tables/test.xlsx")
    message = f"сохранено > test.xlsx"
    flash(message)


def save_info_month(wb_obj_out):
    '''сохраняем в итоговый файл'''
    wb_obj_out.save("temp_tables/test_month.xlsx")
    message = f"сохранено > test_month.xlsx"
    flash(message)


def save_selected_summary_medical_work():
    oracledb.init_oracle_client()
    dsn_tns = oracledb.makedsn('srvmis', '1521', service_name='mntk')
    conn = oracledb.connect(user=r'838', password='1128', dsn=dsn_tns)
    c = conn.cursor()
    query = f"""select cg.id_custom_group, csg.id_subgroup,
           NVL(cg.custom_group_description, '-- БЕЗ ГРУППЫ') GR_NAME,
           csg.subgroup_description,
           msk_opnames.head_oper_description,
           count(*) CNT_OP
    from old_medbase.operzal o

    left outer join old_medbase.titlist tit on (tit.id_card=o.nom_kart)

    left outer join medbase.head_native_cross msk on(o.oper1=msk.native_oper and nvl(o.kod_iol,0)=msk.native_iol)
    left outer join medbase.head_operations msk_opnames on (msk_opnames.head_oper=msk.head_oper)
    left outer join medbase.head_oper_groups gr_msk on (gr_msk.code=msk_opnames.diag_code1)
    --left outer join medbase.kl_values supergr_msk on (gr_msk.head_title=supergr_msk.id_universal)

    left outer join medbase.custom_groups_oper cgo on (cgo.head_oper = msk_opnames.head_oper)
    left outer join medbase.custom_groups cg on (cg.id_custom_group=cgo.id_custom_group)
    left outer join medbase.custom_subgroups csg on (csg.id_subgroup=cgo.id_subgroup)

    left outer join old_medbase.visit v on (o.nom_kart=v.nom_kart and o.nom_pos=v.nom_pos)
    left outer join old_medbase.kl_dogov dogov on (dogov.kod_dogov=v.kod_dogov)
    left outer join old_medbase.kl_regions reg on (reg.id_region=v.kod_regn)

    where o.if_opr=2 and o.oplata=1 and (o.date_olk between TO_DATE('{date1}', 'DD/MM/YY') 
    and TO_DATE('{date2}', 'DD/MM/YY'))
    and (months_between(o.date_olk,tit.birthday)/12 >= {age_from}) 
    and (months_between(o.date_olk,tit.birthday)/12 < {age_to})

    group by cg.id_custom_group, csg.id_subgroup,
             NVL(cg.custom_group_description, '-- БЕЗ ГРУППЫ'),
             csg.subgroup_description,
             msk_opnames.head_oper_description
    order by 1,2
    """
    result = pd.read_sql(query, con=conn)
    result.to_excel("result.xlsx")
    conn.close()


def save_selected_number_of_children():
    oracledb.init_oracle_client()
    dsn_tns = oracledb.makedsn('srvmis', '1521', service_name='mntk')
    conn = oracledb.connect(user=r'838', password='1128', dsn=dsn_tns)
    c = conn.cursor()
    query = f"""select det, count (distinct nom_kart) ak, sum(perv_pos) perv, sum(napr) toOper,sum(ldk) LDK from  
    (
    select 
    nom_kart,case when trunc(months_between(v.date_begv,t.birthday)/12) between 0 
    and 17 then 'Дети' else 'Взрослые' end det,decode(v.nom_pos,1,1,0) perv_pos,case when v.nom_pos=1 
    then decode(v.RES_P,1,1,0) else 0 end napr,
    (select count(*) from old_medbase.opl_serv
     left outer join medbase.kl_serv on (kl_serv.kod_serv=opl_serv.kod_serv)
     where kl_serv.kod_mvyp between 28 and 32 and opl_serv.nom_kart=v.nom_kart and opl_serv.nom_pos=v.nom_pos 
     and v.nom_pos=1 and v.RES_P=1 and rownum<2
     )  ldk
    from old_medbase.visit v
    left outer join old_medbase.titlist t on (t.id_Card=v.nom_kart)
     where (v.date_begv between TO_DATE('{date1}', 'DD.MM.YYYY') and TO_DATE('{date2}', 'DD.MM.YYYY'))
    )
     group by det
     order by 1,2,3
    """
    result = pd.read_sql(query, con=conn)
    result.to_excel("temp_tables/__________2.xlsx")
    conn.close()


def save_selected_summary_flat_data():
    oracledb.init_oracle_client()
    dsn_tns = oracledb.makedsn('srvmis', '1521', service_name='mntk')
    conn = oracledb.connect(user=r'838', password='1128', dsn=dsn_tns)
    c = conn.cursor()
    query = f"""select  
    case 
     when id_uniq_hosp_dir is not null then 'КС'
     when (id_uniq_hosp_dir is null) AND (stday.kdp is not null) then 'ДС'                                           
     else 'Амбулаторно' end "ТипОпер"
    ,
              hog.head_GROUP "ГруппаОперМск",hog.head_group_description "ГруппаОперМскОпис",
              hop.head_oper "ОперМск",ms.head_oper_description "ОперМскОпис",  
             -- o.oper1 "ОперКодКлг", klop.description "ОперКлг",
               kl_cntry.NAME_cntry "СТРАНА",
               KL_REGI.NAME_REGN "Регион (visit.dbf)",
              kld.kod_dogov "КодДоговора",
              kld.name_dogov "Договор",
              k2.description  "ТипДоговора",   
              tit.birthday "Дата рождения",
              trunc(months_between(o.date_olk,tit.birthday)/12,1) "Возраст",
              case when trunc(months_between(o.date_olk,tit.birthday)/12) between 0 
              and 17 then ' Дети' else 'Взрл' end "Д-В",
              decode(tit.sex,1,'М',2,'Ж','Неизвестен') "Пол",
              o.nom_kart "Амб.карта", 
              tit.family_name||' '||tit.name||' '||tit.patronymic "ФИО",
             diag.description "ДиагнозОперированногоГлаза",
             mkb10.descript_mkb "ДиагнозОперированногоГлазаМКБ",
             diag2.description "ДиагнозОперированногоГлаза-2",
             mkb102.descript_mkb "ДиагнозОперированногоМКБ-2",

             nt.description "НТ",
             klh.familia || ' ' || klh.name || ' ' || klh.otchestvo "Хирург",
             o.date_op "Дата операции",
             DECODE(o.eye,1,'OS',2,'OD') "Глаз",
             klbrg.description "Бригада",
             NVL(o.kod_iol,0) "Код линзы",
             MSK_IOLS.descript "Назв линзы",
             TRIM(DECODE(KL_IOL.Soft,'1','мягкая ',NULL)||DECODE(KL_IOL.Aspher,'1','асферическая ',NULL)||DECODE(KL_IOL.Multi,'1','мультифокальная ',NULL)||DECODE(KL_IOL.Toric,'1','торическая ',NULL)||DECODE(KL_IOL.Rear,'1','заднекамерная ',NULL)) "ТипЛинзы",
             lvis.description "Причина снижения зрения",
             o.nprim "Примечание",
             st.DS_DIR_OS "ДиагПриУбытии",st_mkb10.DIAGNOSIS_MKB "ДиагПриУбытии",o.REZULT ,o.kod_c_071, kld.ino_only "INOSTR"

    from old_medbase.operzal o
    left outer join medbase.head_native_cross hop on (hop.native_oper=o.oper1 and hop.native_iol=NVL(o.kod_iol,0))
    Left Outer Join Medbase.Head_Operations Ms on (hop.HEAD_OPER=ms.head_oper)
    left outer join MEDBASE.head_oper_groups hog on (ms.diag_code1=hog.code)
    left outer join MEDBASE.kl_operations klop on (klop.id_operation=o.oper1)
    left outer join KL_IOL@KLBASE on (o.kod_iol=KL_IOL.kod_iol)
    left outer join MEDBASE.MSK_IOLS on (o.kod_iol=MSK_IOLS.kod_iol)
    left outer join MEDBASE.stat_oper_link sol on (sol.id_card=o.nom_kart and sol.num_visit=o.nom_pos and sol.operate_step=o.step)
    inner join OLD_MEDBASE.visit v on (v.nom_kart=o.nom_kart and v.nom_pos=o.nom_pos)
    left outer join old_medbase.kl_dogov kld on (kld.kod_dogov=v.kod_dogov)
    left outer join medbase.kl_values k2 on (k2.value=kld.kod_mvyp and k2.id_value_part=106)
    LEFT OUTER JOIN KL_CNTRy@klbase ON (KL_CNTRy.kod_CNTRY=V.KOD_CNTRY)
    LEFT OUTER JOIN KL_REGI@KLBASE ON (KL_REGI.KOD_REGN=V.KOD_REGN)

    inner join old_medbase.titlist tit on (tit.id_card=o.nom_kart)
    left outer join medbase.kl_diagnoses diag on (diag.id_diagnosis=DECODE(o.eye, 1, v.os_d1, 2, v.od_d1, NULL))
    left outer join medbase.kl_mkb10 mkb10 on (diag.group_diag1=mkb10.group_diag_int)
    left outer join medbase.kl_diagnoses diag2 on (diag2.id_diagnosis=DECODE(o.eye, 1, v.os_d2, 2, v.od_d2, NULL))
    left outer join medbase.kl_mkb10 mkb102 on (diag2.group_diag1=mkb102.group_diag_int)

    LEFT outer join medbase.olk_extra e on (o.nom_kart=e.id_card and o.nom_pos=e.num_visit and o.step=e.step)
    LEFT OUTER join medbase.kl_values nt on (e.nt=nt.id_universal) 
    LEFT OUTER join medbase.kl_values klbrg on (o.kodbrg=klbrg.value and klbrg.id_value_part=103) 
    left outer join kl_hir@KLBASE klh on (klh.kod_hir=o.vrachh)
    left outer join medbase.kl_values lvis on (lvis.value=o.kod_lvis and lvis.id_value_part=34)

    left outer join MEDBASE.stationar st using(id_uniq_hosp_dir)
    left outer join medbase.kl_mkb10 st_mkb10 on (st.DS_DIR_OS=st_mkb10.descript_mkb)


    left outer join OLD_MEDBASE.stday on (o.nom_kart=stday.nom_kart and o.nom_pos=stday.nom_pos)
    where (o.date_olk between TO_DATE('{date1}', 'DD/MM/YY') and TO_DATE('{date2}', 'DD/MM/YY'))
     and o.oplata=1 and o.if_opr=2 
    order by 1,2,4
    """
    result = pd.read_sql(query, con=conn)
    result.to_excel("temp_tables/__________.xlsx")
    conn.close()


def save_selected_summary_sources_fin():
    oracledb.init_oracle_client()
    dsn_tns = oracledb.makedsn('srvmis', '1521', service_name='mntk')
    conn = oracledb.connect(user=r'838', password='1128', dsn=dsn_tns)
    c = conn.cursor()
    query = f"""select  
    case 
     when sol.id_uniq_hosp_dir is not null then 'КС'
     when (id_uniq_hosp_dir is null) AND (stday.kdp is not null) then 'ДС'                                           
     else 'Амбулаторно' end "АмбСтац",
            case 
                 when months_between(o.date_olk,tit.birthday)/12 < 18 then 'Дети' 
                 when months_between(o.date_olk,tit.birthday)/12 >= 18 then 'Взрослые' 
                 else 'Нет данных' 
              end "Возрастная категория",
              klv.description "Тип договора",                              
              kld.name_dogov "Договор",
              decode(og.head_group, 'C', 'Конcерв', 'Хирургия') "Вид лечения", 
              count(distinct o.nom_kart||'-'||o.nom_pos) "Всего уник. амб.карт+пос"

    from old_medbase.operzal o
    left outer join MEDBASE.stat_oper_link sol on (sol.id_card=o.nom_kart and sol.num_visit=o.nom_pos and sol.operate_step=o.step)
    left outer join old_medbase.titlist tit on (tit.id_card=o.nom_kart)
    --left outer join medbase.kladr on (tit.kod_c||'00000000000'=kladr.code)
    left outer join OLD_MEDBASE.visit v on (v.nom_kart=o.nom_kart and v.nom_pos=o.nom_pos)
    left outer join old_medbase.kl_dogov kld on (kld.kod_dogov=v.kod_dogov)
    left outer join medbase.kl_values klv on (klv.id_value_part=106 and klv.value=kld.kod_mvyp)

    LEFT OUTER JOIN medbase.head_native_cross msk ON (o.oper1=msk.native_oper AND NVL(o.kod_iol,0)=msk.native_iol  )
    LEFT OUTER JOIN medbase.head_operations using(head_oper)
    left outer join medbase.head_oper_groups og on (og.code=medbase.head_operations.diag_code1)
    left outer join OLD_MEDBASE.stday on (o.nom_kart=stday.nom_kart and o.nom_pos=stday.nom_pos)
    where o.date_olk between TO_DATE('{date1}', 'DD/MM/YY') and TO_DATE('{date2}', 'DD/MM/YY')
     and o.oplata=1 and o.if_opr=2                          

     and 
     (og.head_group<>'C'  -- хирургия
              or          --    или
     (og.head_group='C' and not exists (
                             select * from old_medbase.operzal subo 
                             LEFT OUTER JOIN medbase.head_native_cross submsk ON (subo.oper1=submsk.native_oper 
                             AND NVL(subo.kod_iol,0)=submsk.native_iol  )
                             LEFT OUTER JOIN medbase.head_operations subhead_operations using(head_oper)
                             left outer join medbase.head_oper_groups subog on (subog.code=subhead_operations.diag_code1)
                             where subo.nom_kart=o.nom_kart and subo.nom_pos=o.nom_pos and subo.step<>o.step 
                             and subog.head_group<>'C' 
                             and subo.date_olk between TO_DATE('{date1}', 'DD/MM/YY') and TO_DATE('{date2}', 'DD/MM/YY')
     )  )
     )
    group by case 
     when sol.id_uniq_hosp_dir is not null then 'КС'
     when (id_uniq_hosp_dir is null) AND (stday.kdp is not null) then 'ДС'                                           
     else 'Амбулаторно' end ,
            case 
                 when months_between(o.date_olk,tit.birthday)/12 < 18 then 'Дети' 
                 when months_between(o.date_olk,tit.birthday)/12 >= 18 then 'Взрослые' 
                 else 'Нет данных' 
              end,
              klv.description, kld.name_dogov,
              decode(og.head_group, 'C', 'Конcерв', 'Хирургия')

    order by 1,2,3,4
    """
    result = pd.read_sql(query, con=conn)
    result.to_excel("temp_tables/__________1.xlsx")
    conn.close()


def save_selected_summary_vmp():
    oracledb.init_oracle_client()
    dsn_tns = oracledb.makedsn('srvmis', '1521', service_name='mntk')
    conn = oracledb.connect(user=r'838', password='1128', dsn=dsn_tns)
    c = conn.cursor()
    query = f"""select 
    vmp.vmp_part,vmp.vmp_group,
    substr(vmp.i10,1,instr(vmp.i10,';')-1) Region, visit.smo SMO_visit,  kl_smorg.desc_reg SMO_REGION_visit,
    vmp.id_card NOM_KART,-- o.nom_pos, 
       stationar.officiAL_NO||'-'||to_char(stationar.date_out,'YYYY') StacNo,
    titlist.family_name||' '|| titlist.name||' '|| titlist.patronymic FAMILIA, -- ФИО
    SUBSTR(vmp.talon_no,1,2)||'.'||SUBSTR(vmp.talon_no,3,4)||'.'||SUBSTR(vmp.talon_no,7,5)||'.'||SUBSTR(vmp.talon_no,12,3) TALONE, --Талон ВМП
    TO_CHAR(titlist.birthday,'DD.MM.YYYY') AGE, --ДР
              case 
                 when months_between(stationar.date_out,titlist.birthday)/12 < 18 then 'Дети' 
                 when months_between(stationar.date_out,titlist.birthday)/12 >= 18 then 'Взрослые' 
                 else 'Нет данных' 
              end "Возрастная категория",
    --translate(kl_mkb10.descript_mkb,'~+','_') MKB, --Код диагноза МКБ10 (убирается +)
    medbase.stationar.DS_DIR_OS,
    nvl2(Vmp.mu5_6,'11.00.'||vmp_standards.printable_description,Vmp.mu5_6) Gr_Ht,
    --nvl2(Vmp.mu5_6,substr(Vmp.mu5_6,1,3)||'00'||'.00'||substr(Vmp.mu5_6,4,1),Vmp.mu5_6) Gr_Ht,

    TO_CHAR(stationar.date_in,'DD.MM.YYYY HH24:MI') date_in,
    TO_CHAR(stationar.date_out,'DD.MM.YYYY HH24:MI') date_out,
    Kind.Description kind
    , methods.description method        ,
    finals.description result,
    DECODE(medbase.is_diagnos_vmp_correct(vmp.mu5_6,medbase.stationar.DS_DIR_OS),1,'Да','Нет') IsDiagCorrect


    from old_medbase.operzal o
    left outer join old_medbase.visit on (visit.nom_kart= o.nom_kart and visit.nom_pos= o.nom_pos)
    left outer join old_medbase.kl_smorg on (substr(visit.smo,1,2)=kl_smorg.id_reg)
    left outer join old_medbase.titlist on (titlist.id_card= o.nom_kart)
    left outer join medbase.vmp_talons vmp on (vmp.id_card= o.nom_kart and vmp.visit_no= o.nom_pos and vmp.talon_state=1034)
    left outer join medbase.vmp_standards on (vmp_standards.id_vmp_standard=vmp.mu5_6)
    --left outer join medbase.kl_diagnoses on (DECODE (o.eye,1,visit.os_d1,2,visit.od_d1)=kl_diagnoses.id_diagnosis)
    --left outer join medbase.kl_mkb10 on (kl_diagnoses.group_diag1=kl_mkb10.group_diag_int)
    -- Стационар
    LEFT OUTER JOIN medbase.stat_oper_link sol on (sol.id_card=o.nom_kart AND sol.num_visit=o.nom_pos AND sol.operate_step=o.step)
    LEFT OUTER JOIN medbase.stationar using (id_uniq_hosp_dir)
    LEFT OUTER JOIN medbase.kl_values finals on (finals.id_value_part=59 AND finals.id_universal=stationar.ID_HOSP_RESULT)
    --Left Outer Join Old_Medbase.Pop_Osl On (Pop_Osl.Nom_Kart=O.Nom_Kart And Pop_Osl.Nom_Pos=O.Nom_Pos And Pop_Osl.Step=O.Step)
    Left Outer Join Medbase.Kl_Values Kind On (Kind.Id_Universal=Vmp.Id_Kind)
    Left Outer Join Medbase.vmp_methods methods On (methods.Id_method=Vmp.Id_method)

    Where 
    o.oplata=1 -- and   o.nom_kart=343218
    and o.if_opr=2 
    and
    visit.kod_dogov in (155,199,200,201,209,210,216,231,269,270) and -- Только ВТ

    stationar.date_out between TO_DATE('{date_begin_year}', 'DD/MM/YY') and TO_DATE('{date2}', 'DD/MM/YY')

    group by vmp.vmp_part,vmp.vmp_group,vmp.i10,visit.smo,    kl_smorg.desc_reg,
    vmp.id_card,        stationar.officiAL_NO||'-'||to_char(stationar.date_out,'YYYY'),
    titlist.family_name||' '|| titlist.name||' '|| titlist.patronymic,
    SUBSTR(vmp.talon_no,1,2)||'.'||SUBSTR(vmp.talon_no,3,4)||'.'||SUBSTR(vmp.talon_no,7,5)||'.'||SUBSTR(vmp.talon_no,12,3) ,
    TO_CHAR(titlist.birthday,'DD.MM.YYYY'),
              case 
                 when months_between(stationar.date_out,titlist.birthday)/12 < 18 then 'Дети' 
                 when months_between(stationar.date_out,titlist.birthday)/12 >= 18 then 'Взрослые' 
                 else 'Нет данных' 
              end,
    --translate(kl_mkb10.descript_mkb,'~+','_'),
    medbase.stationar.DS_DIR_OS,
    nvl2(Vmp.mu5_6,'11.00.'||vmp_standards.printable_description,Vmp.mu5_6),
    TO_CHAR(stationar.date_in,'DD.MM.YYYY HH24:MI'),
    TO_CHAR(stationar.date_out,'DD.MM.YYYY HH24:MI'),
    Kind.Description, methods.description ,finals.description,
    DECODE(medbase.is_diagnos_vmp_correct(vmp.mu5_6,medbase.stationar.DS_DIR_OS),1,'Да','Нет')
    order by 1,2,4
    """
    result = pd.read_sql(query, con=conn)
    result.to_excel("temp_tables/vmp.xlsx")
    conn.close()





list_for_nums_of_rows_konserv, list_for_nums_of_rows_deti, list_for_nums_of_rows_vmp, \
list_for_nums_of_rows_deti_vmp, list_for_nums_hk, list_for_nums_iol_1, list_for_nums_end, list_for_nums_25g_1, \
list_for_nums_end_2, list_for_nums_27g_1, list_for_nums_vhk, list_for_nums_25g_2, list_for_nums_vhhk, \
list_for_nums_27g_2, list_for_nums_of_rows_kk, all_card_kids, list_for_nums_oper_deti, \
list_for_nums_pleoptika, list_foreign_peoples, list_for_nums_of_rows_c006_kids = ([] for _ in range(20))


reporting_month = {'01':'Январь', '02':'Февраль', '03':'Март', '04':'Апрель', '05':'Май', '06':'Июнь', '07':'Июль', '08':'Август',
                   '09':'Сентябрь', '10':'Октябрь', '11':'Ноябрь', '12':'Декабрь'}


path_from_old_year = "temp_tables/__________old_year.xlsx"
path_sources_old_year = "temp_tables/__________1_old_year.xlsx"
path_kids_old_year = "temp_tables/__________2_old_year.xlsx"
path_vmp_old_year = 'temp_tables/vmp_old_year.xlsx'

path_from = "temp_tables/__________.xlsx"
path_sources = "temp_tables/__________1.xlsx"
path_kids = "temp_tables/__________2.xlsx"
path_vmp = 'temp_tables/vmp.xlsx'

#path_output = "RESULT_flat.xlsx"
#path_output = "\\\\192.168.0.100\\disku\\DIANA\\test.xlsx"
path_output_weekly = "temp_tables/test.xlsx"
path_output_monthly = "temp_tables/test_month.xlsx"